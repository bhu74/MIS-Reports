"""Helper functions"""

from math import ceil, isnan
import re
from datetime import datetime
from decimal import Decimal

from pandas import DataFrame, Series
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows  # pylint:disable=unused-import
from loguru import logger
import numpy as np

import config as cfg
import src.excel_helper as excel_helper


class MissingValueError(Exception):
    """
    Custom exception for errors when desired value is missing
    """


def read_sheet(file_name, sheet_names, is_header_present=False, is_read_only=False, is_data_only=True):
    """
    Function to read the excel sheet
    Parameters:
        file_name - Excel file name
        sheet_names - One sheet or a list of sheets which need to be read from excel file
        is_header_present - Is the first row column of the data
        is_read_only - Should the file be opened in read only mode
    Returns:
        Data frame with values read from sheet
    """
    data_dict = {}
    workbook = load_workbook(file_name, read_only=is_read_only, data_only=is_data_only)

    if isinstance(sheet_names, str):
        if sheet_names.lower() == 'all':
            sheet_names = workbook.sheetnames
        else:
            sheet_names = [sheet_names]

    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            logger.error("Sheet {} not found in {}".format(sheet_name, file_name))
            exit(-1)
        data = DataFrame(workbook[sheet_name].values)
        if is_header_present:
            headers = data.iloc[0]
            data = data[1:]
            data.rename(columns=headers, inplace=True)
        data_dict[sheet_name] = data

    if len(sheet_names) == 1:
        return data_dict[sheet_names[0]]

    return data_dict


def add_metadata(data_frame):
    """
    Add metadata rows and columns in the given data frame
    Parameters:
        data_frame (Pandas dataframe)
    Returns:
        Data frame with metadata row (ar) and column (ac)
    """
    data_frame.fillna('', inplace=True)
    data_frame = data_frame.append(Series(data_frame.apply(
        lambda value: '{1}{0}{1}'.format('|'.join(map(str, value)), '|'), axis=0), name='ar'))
    data_frame['ac'] = data_frame.apply(
        lambda value: '{1}{0}{1}'.format('|'.join(map(str, value)), '|'), axis=1)
    data_frame.at[len(data_frame.index)-1, 'ac'] = ''
    return data_frame


def strip_metadata(data_frame):
    """
    Remove metadata rows and columns in the given data frame
    Parameters:
        data_frame (Pandas dataframe)
    Returns:
        Data frame with metadata row (ar) and column (ac) removed
    """
    # Drop aggregate row and column
    data_frame.drop(['ar'], axis=0, inplace=True)
    data_frame.drop(['ac'], axis=1, inplace=True)
    data_frame = data_frame.replace(r'^\s*$', None, regex=True)
    return data_frame


def get_row_index(data_frame, row_cond, row_start=0):
    """
    Find index of row in dataframe that contains specified condition
    Parameters:
        data_frame (Pandas DataFrame)
        row_cond (String)
        row_start (Int) - index of row where the search should begin from
    Returns:
        Index of row that contains specified condition
    """
    if row_cond:
        return data_frame.iloc[row_start:][
            data_frame.iloc[row_start:]['ac'].str.contains(
                str(row_cond), regex=True)].first_valid_index()
    return 0


def get_col_index(data_frame, col_cond, col_start=0, row_start=0):
    """
    Find index of column in dataframe that contains specified condition
    Parameters:
        data_frame (Pandas DataFrame)
        col_cond (String)
        col_start (Int) - index of column where the search should begin from
        row_start (Int) - index of row where the search should begin from
    Returns:
        Index of column that contains specified condition
    """
    if col_cond:
        filtered_cols = data_frame.iloc[row_start:, col_start:].loc['ar']\
            .str.contains(str(col_cond), regex=True)
        return filtered_cols[filtered_cols].first_valid_index()
    return 0


# pylint: disable=too-many-locals, too-many-statements
def append_suffix(alias_name, suffix=""):
    """
    Add the tab suffix to the input alias name
    Parameters:
        alias_name (String)
        suffix (String)
    Returns:
        New alias name with suffix added
    """
    if alias_name[0] == 'r' or alias_name[0] == 'c':
        if len(alias_name.split('+')) > 1:
            alias_new = alias_name.split('+')[0] + suffix + '+' + alias_name.split('+')[1]
        elif len(alias_name.split('-')) > 1:
            alias_new = alias_name.split('-')[0] + suffix + '-' + alias_name.split('-')[1]
        else:
            alias_new = alias_name+suffix
    else:
        alias_new = alias_name
    return alias_new


# pylint: disable=too-many-locals, too-many-statements
def replace_alias(statement, old_alias_list, new_alias_list):
    """
    Replace the alias names in old_alias_list with alias names in new_alias_list
        in the input statement
    Parameters:
        statement (String)
        old_alias_list (String)
        new_alias_list (String)
    Returns:
        Statement with alias names replaced with alias names in new_alias_list
    """
    statement_halfs = ['', statement]
    new_statement = ''
    for old_alias, new_alias in zip(old_alias_list, new_alias_list):
        statement_halfs = statement_halfs[1].split(old_alias, 1)
        new_statement = new_statement + statement_halfs[0] + new_alias

    try:
        new_statement = new_statement + statement_halfs[1]
    except:  # pylint: disable=bare-except
        new_statement = statement

    return new_statement


# pylint: disable=too-many-locals, too-many-statements
def add_suffix(statement):
    """
    Add suffix to the alias names in the given statement
    Parameters:
        statement (String)
    Returns:
        Statement with suffix added to alias names
    """
    update_list = []
    suffix_list = re.findall(r"([a-z_]+)(_source|_dest|_tab1|_tab2)", statement)
    if suffix_list:
        alias_block_list = statement.split('][')
        alias_list = [alias_block_list[0][alias_block_list[0].rfind('[')+1:]]
        for alias_block in alias_block_list[1:-1]:
            alias_left = alias_block[:alias_block.find(']')]
            alias_right = alias_block[alias_block.rfind('[')+1:]
            alias_list = alias_list + [alias_left, alias_right]
        alias_list = alias_list + [alias_block_list[-1][:alias_block_list[-1].find(']')]]
        alias_list = [pair.split(':') for pair in alias_list]
        alias_list = [item for sublist in alias_list for item in sublist]

        for index, suffix_pair in enumerate(suffix_list):
            suffix = suffix_pair[0]
            for a_name in alias_list[index*2:index*2+2]:
                update_list.append(append_suffix(a_name, suffix))
        statement = replace_alias(statement, alias_list, update_list)
    return statement


def apply_statement(statement, apply_rows=1, apply_cols=1):
    """
    Apply statement to given rows and columns
    Parameters:
        statement {List of statements}
        apply_rows - number of rows to apply statement on
        apply_cols - number of columns to apply statement on
    Returns:
        Numpy array of derived statements at the corresponding row, column
    """
    stmt_lst = []
    if statement[0] == "[":
        res = statement.strip('][').split("\"")
        for r_no in res:
            if len(r_no) > 3:
                stmt_lst.append(r_no)
    else:
        stmt_lst.append(statement)

    x_f1 = lambda x: len(stmt_lst) if x == 1 else x
    ret_stmt = np.empty([apply_rows, x_f1(apply_cols)], dtype=object)

    if apply_rows > 1 and apply_cols > 1:
        first_stmt = stmt_lst[0]
        stmt_lst.clear()
        for j in range(apply_cols):
            next_stmt = get_next_statement(first_stmt, 'col', j)
            stmt_lst.append(next_stmt)
        apply_cols = 1

    for c_num, s_name in enumerate(stmt_lst):
        stmt_suffix = add_suffix(s_name)
        for i in range(apply_rows):
            next_stmt = get_next_statement(stmt_suffix, 'row', i)
            ret_stmt[i][c_num] = next_stmt

        if apply_cols > 1:
            for j in range(apply_cols):
                next_stmt = get_next_statement(stmt_suffix, 'col', j)
                ret_stmt[c_num][j] = next_stmt
    return ret_stmt


# pylint: disable=too-many-locals, too-many-statements
def get_next_statement(statement, apply_to, cell_num):
    """
    Derives the next statement for the given row/column at the cell number
    Parameters:
        statement (String)
        apply_to (String) - row/col
        cell_num (Integer)
    Returns:
        Next Statement to be applied at the row/col cell number
    """
    row_lst = []
    col_lst = []

    alias_lst = re.findall(r"\[(.[\w\+\-]+?)\]", statement)
    for a_name in alias_lst:
        if a_name[0] == 'r':
            row_lst.append(a_name)
        elif a_name[0] == 'c':
            col_lst.append(a_name)

    new_rows = row_lst.copy()
    new_cols = col_lst.copy()

    if apply_to == 'row':
        for i, _ in enumerate(row_lst):
            if len(row_lst[i].split('+')) > 1:
                off_set = int(row_lst[i].split('+')[1]) + cell_num
                new_rows[i] = row_lst[i].split('+')[0] + '+' + str(off_set)
            else:
                off_set = cell_num
                new_rows[i] = row_lst[i] + '+' + str(off_set)

    if apply_to == 'col':
        for j, _ in enumerate(col_lst):
            if len(col_lst[j].split('+')) > 1:
                off_set = int(col_lst[j].split('+')[1]) + cell_num
                new_cols[j] = col_lst[j].split('+')[0] + '+' + str(off_set)
            else:
                off_set = cell_num
                new_cols[j] = col_lst[j] + '+' + str(off_set)

    statement = replace_alias(statement, row_lst, new_rows)
    statement = replace_alias(statement, col_lst, new_cols)

    return statement


# pylint: disable=too-many-locals, too-many-statements
def calc_month_table(report_month):
    """
    Routine to populate the month table of Input Tab
    Parameters:
        Report Month {Integer}
    Returns:
        Numpy array of month table values
    """

    table_rows = 21
    table_cols = 10
    current_mth_row = 7
    ret_table = np.empty([table_rows, table_cols], dtype=object)

    tbl_mth = report_month.month
    rpt_yr = report_month.year

    for i in range(current_mth_row, -1, -1):
        ret_table[i][0] = tbl_mth
        ret_table[i][1] = month_name(tbl_mth)
        ret_table[i][2] = month_long_name(tbl_mth)
        ret_table[i][3] = "Q"+str(ceil(tbl_mth/3))
        ret_table[i][4] = rpt_yr
        ret_table[i][5] = "Y"+str(rpt_yr)
        if tbl_mth == 12:
            tbl_mth = 1
            rpt_yr += 1
        else:
            tbl_mth += 1

    tbl_mth = report_month.month
    rpt_yr = report_month.year
    for j in range(current_mth_row+1, table_rows, 1):
        if tbl_mth == 1:
            tbl_mth = 12
            rpt_yr -= 1
        else:
            tbl_mth -= 1
        ret_table[j][0] = tbl_mth
        ret_table[j][1] = month_name(tbl_mth)
        ret_table[j][2] = month_long_name(tbl_mth)
        ret_table[j][3] = "Q"+str(ceil(tbl_mth/3))
        ret_table[j][4] = rpt_yr
        ret_table[j][5] = "Y"+str(rpt_yr)

    qtr_row = ceil(ret_table[0][0]/3)
    for k in range(table_rows):
        ret_table[k][6] = qtr_row
        ret_table[k][7] = str((ret_table[k][3]))[1:]+"Q"+str((ret_table[k][4]))[2:]
        ret_table[k][8] = str(ret_table[k][7]) + str(ret_table[k][6])
        ret_table[k][9] = ret_table[k][1]
        if qtr_row == 1:
            qtr_row = 3
        else:
            qtr_row -= 1

    return ret_table


# pylint: disable=too-many-locals, too-many-statements
def cell_diff(a_num, b_num):
    """
    Calculates difference between two cell values (handles integers/integers stored as strings)
    Parameters:
        a_num {Integer/Integer passed as String} - First number
        b_num {Integer/Integer passed as String} - Second number
    Returns:
        {Integer} - Difference
    """
    try:
        a_dec = Decimal(a_num)
    except:  # pylint: disable=bare-except
        a_dec = 0
    try:
        b_dec = Decimal(b_num)
    except:  # pylint: disable=bare-except
        b_dec = 0

    result = a_dec-b_dec
    if isnan(result):
        return 0

    return result


# pylint: disable=too-many-locals, too-many-statements, no-else-return
def cell_div(f_val, s_val):
    """
    Divides two values (handles integers/integers stored as strings) \
        and returns result or error message
    Parameters:
        f_val {Integer/Integer passed as String} - First number
        s_val {Integer/Integer passed as String} - Second number
    Returns:
        {Integer} - Result
    """
    try:
        return Decimal(f_val)/Decimal(s_val)
    except:         # pylint: disable=bare-except
        if str(s_val) == "0":
            return "#DIV/0!"
        else:
            return '#VALUE!'


def month_long_name(num):
    """
    Get full month name based on the month number
    Parameters:
        num {Integer} - Number of the month
    Returns:
        {String} - Shortened month name for the given month number
    """
    return 'January February March April May June July August September October \
            November December'.split()[num - 1]


# pylint: disable=too-many-locals, too-many-statements
def get_prev_mth(date):
    """
    Get the previous month of the date passed
    Parameters:
        Date {Date} - Input Date
    Returns:
        {Date} - Date of previous month
    """
    if date.month == 1:
        prev_mth = date.replace(month=12, year=date.year-1)
    else:
        prev_mth = date.replace(day=1, month=date.month-1)
    return prev_mth


# pylint: disable=too-many-locals, too-many-statements, no-else-return
def cell_sum(cells):
    """
    Returns the sum of a list of values (handles integers/integers stored as strings) \
        and returns result or error message
    Parameters:
        cells {list} - List of values
    Returns:
        {Integer} - Result
    """
    hasDiv0 = False
    hasError = False

    lst = []
    for i in cells:
        check_err = []
        text = str(i).strip()
        check_err = re.findall(r"[!#a-zA-z]+", text.lower())
        if check_err and ('e' not in check_err) and i is not None:
            hasError = True
            if text == "#DIV/0!":
                hasDiv0 = True
        else:
            try:
                i = float(i)
            except:         #pylint: disable=bare-except
                i = 0.0
            lst.append(i)

    if hasDiv0:
        return "#DIV/0!"
    if hasError:
        return "#VALUE!"

    return Decimal(sum(lst))

# pylint: disable=too-many-locals, too-many-statements, no-else-return
def check_grouping(cells):
    """
    Returns grouping check result.
    Parameters:
        cells {list} - List of values
    Returns:
        {String} - Result
    """
    s = cell_sum(cells)
    if s == "#VALUE!":
        return s
    s = abs(s)
    if s < 100000:
        return '1'
    else:
        return '0'

# pylint: disable=too-many-locals, too-many-statements
def float_val(num):
    """
    Returns the float value of a number stored as string
    Parameters:
        num {String or Integer} - Value
    Returns:
        {Float} - Result
    """
    try:
        return float(num)
    except:   # pylint: disable=bare-except
        return 0


def negative_value(num):
    """
    Return negative value from input value
    Parameters:
        num {Integer/Float} - Input value
    Returns:
        {Integer/Float} - Negative value
    """
    if num == '':
        raise MissingValueError
    if isinstance(num, (int, float)):
        return -num
    elif num == '0':
        return 0
    else:
        raise ValueError


# pylint: disable=too-many-locals, too-many-statements, too-many-nested-blocks, no-else-return
def calcPercentage(first_val, second_val):
    """
    Calculate percentage as per the formula in Country Financials reports
    Parameters:
        first_val {Integer} - First number
        second_val {Integer} - Second number
    Returns:
        {Integer/String} - Calculated value
    """
    # pylint: disable-msg=too-many-return-statements
    if (first_val == 'n/m') or (second_val == 'n/m'):
        return "#VALUE!"
    elif not str(first_val).strip() and not str(second_val).strip():
        return "n/m"
    elif first_val is None or not str(first_val).strip():
        first_val = 0.0
    elif second_val is None or not str(second_val).strip():
        second_val = 0.0

    check_err1 = re.findall(r"[!#a-zA-z]+", str(first_val).strip().lower())
    check_err2 = re.findall(r"[!#a-zA-z]+", str(second_val).strip().lower())
    if check_err1 and ('e' not in check_err1):
        return "n/m"
    if check_err2 and ('e' not in check_err2):
        return "n/m"

    try:
        first_val = Decimal(first_val)
        second_val = Decimal(second_val)
        calc_val = 0
        if second_val == 0:
            return 'n/m'
        else:
            calc_val = (first_val - second_val)/second_val * 100
            if second_val < 0 and (calc_val < -500):
                return '>500'
            else:
                if second_val < 0 and calc_val > 500:
                    return '<-500'
                else:
                    if second_val < 0:
                        return calc_val * -1
                    else:
                        if calc_val > 500:
                            return '>500'
                        else:
                            if calc_val < -500:
                                return '<-500'
                            else:
                                return calc_val
    except:             # pylint: disable=bare-except
        return 0


# pylint: disable=too-many-locals, too-many-statements
def div_check(f_val, s_val):
    """
    Divide 2 numbers and return result or error code 'n/m'
    Parameters:
        f_val {String or Integer} - First number
        s_val {String or Integer} - Second number
    Returns:
        {Integer} - Result
    """
    try:
        return Decimal(f_val)/Decimal(s_val)
    except:             # pylint: disable=bare-except
        return 'n/m'


# pylint: disable=too-many-locals, too-many-statements
def ci_ratio(f_val, s_val):
    """
    calculate Ci_ratio between 2 numbers and return result or error code 'n/m'
    Parameters:
        f_val {String or Integer} - First number
        s_val {String or Integer} - Second number
    Returns:
        {Integer} - Result
    """
    try:
        f_val = Decimal(f_val)
        s_val = Decimal(s_val)
        if s_val == 0:
            return "#DIV/0!"
        ret = 0
        if f_val * 100/s_val > 500:
            return "n/m"
        else:
            if f_val * 100/s_val < -500:
                return "n/m"
            else:
                return f_val/s_val
    except:             # pylint: disable=bare-except
        return 'n/m'


# pylint: disable=too-many-branches
def check_alias_row(row, idx, file, sheet_df):
    """
    Check validity of alias row and return error if the alias is not valid
    Parameters:
        row - {Pandas DataSeries}
        idx - row index in the alias file
        file - name of the alias file
        sheet_df - {Pandas DataFrame}
    """
    if row['Alias'][0] == 'r':
        start_row = get_row_index(sheet_df, row['start row/col'])
        if start_row is None:
            logger.warning(cfg.INVALID_ALIAS_MESSAGE.format(row['start row/col'], idx + 2, file))
            return None
        if row['Keyword'] == '':
            logger.error(cfg.INVALID_ALIAS_MESSAGE.format(row['Keyword'], idx + 2, file))
            exit(-1)
        try:
            row_index = get_row_index(sheet_df, row['Keyword'], start_row)
        except NameError:
            logger.error(cfg.INVALID_ALIAS_MESSAGE.format(row['Alias'], idx + 2, file))
            exit(-1)
        except TypeError:
            logger.error(cfg.INVALID_ALIAS_MESSAGE.format(row['Keyword'], idx + 2, file))
            exit(-1)
        if row_index is None or row_index == 'ar':
            logger.warning(cfg.INVALID_ALIAS_MESSAGE.format(row['Alias'], idx + 2, file))
            return None

    elif row['Alias'][0] == 'c':
        start_col = get_col_index(sheet_df, row['start row/col'])
        if start_col is None:
            logger.warning(cfg.INVALID_ALIAS_MESSAGE.format(
                row['start row/col'], idx + 2, file))
            return None
        if row['Keyword'] == '':
            logger.error(cfg.INVALID_ALIAS_MESSAGE.format(row['Keyword'], idx + 2, file))
            exit(-1)
        try:
            col_index = get_col_index(sheet_df, row['Keyword'], start_col)
        except NameError:
            logger.error(cfg.INVALID_ALIAS_MESSAGE.format(row['Alias'], idx + 2, file))
            exit(-1)
        except TypeError:
            logger.error(cfg.INVALID_ALIAS_MESSAGE.format(row['Keyword'], idx + 2, file))
            exit(-1)
        if col_index is None or col_index == 'ac':
            logger.warning(cfg.INVALID_ALIAS_MESSAGE.format(row['Alias'], idx + 2, file))
            return None


def rows_to_sum(data_frame, column, row_start, row_end):  # todo docstring
    val_lst = []
    start = get_row_index(data_frame, row_start)
    end = get_row_index(data_frame, row_end) + 1
    col_idx = get_col_index(data_frame, column)
    for row_idx in range(start, end):
        val_lst.append(data_frame.at[row_idx, col_idx])
    return val_lst


# pylint: disable-msg=too-many-arguments
def lookup(data_frame, row_cond, col_cond, row_offset=0, col_offset=0, row_start=0,
           col_start=0, cast_to_float=True):
    """
    Lookup values in the given data frame based on give row and column conditions
    Parameters:
        data_frame {Pandas dataframe}
        row_cond - Condition to be used to filter row
        col_cond - Condition to be used to filter column
        row_offset - Offset of the actual row from the filtered row index
        col_offset - Offset of the actual column from the filtered column index
        row_start - Index of the starting row
        col_start - Index of the starting col
        cast_to_float - Convert looked up value to float
    Returns:
        Looked up value from data frame based on the conditions
    """
    row_index = get_row_index(data_frame, row_cond, row_start)
    col_index = get_col_index(data_frame, col_cond, col_start)
    try:
        result = data_frame.at[row_index + row_offset, col_index + col_offset]
    except TypeError:
        result = 0

    if cast_to_float:
        try:
            return Decimal(result)
        except:
            return 0

    return result


def get_quarter(date):
    """
    Get Quarter number for the given date
    Parameters:
        date {Date} - Input date
    Returns:
        {Integer} - Quarter number of the given date
    """
    return ceil(date.month/3)


def month_name(num):
    """
    Get shortened month name based on the month number
    Parameters:
        num {Integer} - Number of the month
    Returns:
        {String} - Shortened month name for the given month number
    """
    return 'Jan Feb Mar Apr May Jun Jul Aug Sep Oct Nov Dec'.split()[num - 1]


def month_number(name):
    """
    Get month number based on the shortened month name
    Parameters:
        name {String} -Shortened month name
    Returns:
        {String} - Month number
    """
    months = 'JAN FEB MAR APR MAY JUN JUL AUG SEP OCT NOV DEC'.split()
    if name.upper() in months:
        return months.index(name.upper()) + 1

    return 0


def calculate_row_sum(data_frame, start_row_id, end_row_id, col_id):
    """
    Function to calculate sum of a column in the range of rows
    Arguments:
        data_frame {DataFrame} -- DataFrame with actual values
        start_row_id {Integer} -- Index of the starting row
        end_row_id {Integer} -- Index of the ending row
        col_id {Integer} -- Index of the column which need to be summed up
    Returns:
        Integer -- Aggregated value
    """
    start_row = data_frame[data_frame['ac'].str.contains(start_row_id, regex=True)] \
                                                        .first_valid_index()
    end_row = data_frame[data_frame['ac'].str.contains(end_row_id, regex=True)] \
                                                        .first_valid_index()
    filtered_cols = data_frame.loc['ar'].str.contains(col_id, regex=True)
    col_index = filtered_cols[filtered_cols].first_valid_index()
    # Filter empty strings from the result to get proper aggregated value
    vals = filter(None, data_frame.iloc[start_row:end_row, \
                    col_index:(col_index + 1)][col_index].values.tolist())
    try:
        return sum([Decimal(x) for x in vals])
    except ValueError:
        return sum(vals)


def clear_formulae(country):
    """
    Function to clear formulae in output file
    Arguments:
        country {String} -- Country
    Returns:
        Nil
    """
    cob_date = datetime.strptime(cfg.COUNTRY_DATE, '%d-%b-%Y')
    prev_month = get_prev_mth(cob_date)
    country_input_file = cfg.INPUT_DIR + \
                            cfg.INPUT_COUNTRY_FILE.format(prev_month.strftime("%b'%y"), country)
    country_report_file = cfg.OUTPUT_DIR + cfg.OUTPUT_FILE_FORMAT.format(prev_month.strftime("%b'%y"), country)
    EXT_DIC = excel_helper.extract_worksheet_extlst(country_input_file)

    country_report = load_workbook(country_report_file)
    for report_ws in country_report.sheetnames:
        output_sheet = country_report[report_ws]
        for i in range(1, output_sheet.max_row+1):
            for j in range(1, output_sheet.max_column+1):
                cell = output_sheet.cell(row=i, column=j).value
                if type(cell).__name__ != 'MergedCell':
                    if cell in [None, '']:
                        continue
                    elif str(cell)[0] == '=':
                        output_sheet.cell(row=i, column=j).value = None

    country_report.save(country_report_file)
    excel_helper.add_extlst_element(country_report_file, EXT_DIC)

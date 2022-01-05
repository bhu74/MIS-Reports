""" Generate Country Financials reports from the provided input files """

from datetime import datetime
from math import floor # pylint: disable=unused-import
from os import listdir
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from loguru import logger
from tqdm.auto import tqdm
import pandas as pd
import numpy as np
from src.helper import *  # pylint: disable=wildcard-import, unused-wildcard-import
import config as cfg


# pylint: disable=too-many-locals, eval-used, too-many-statements, too-many-nested-blocks, too-many-branches
def generate_country_exp_report(country, country_input_data, country_report_data,
                                country_report, suffix):
    """
    Function to generate EXP report from given input files
    """
    cob_date = datetime.strptime(cfg.COUNTRY_DATE, '%d-%b-%Y')
    prev_month = get_prev_mth(cob_date)

    # Load the worksheets into memory
    exp_sheet = country_report["Exp"]
    input_source = add_metadata(country_report_data['Input'])
    exp_source = add_metadata(country_report_data['Exp'])
    pb_source = add_metadata(country_report_data['PB'])
    ibcm_source = add_metadata(country_report_data['IBCM'])
    mkts_source = add_metadata(country_report_data['Mkts'])
    afg_source = add_metadata(country_report_data['AFG'])

    if country in cfg.GROUP1_COUNTRIES:
        input_mapping_file = cfg.MAPPING_EXP_TAB.format('group1')
        input_mapping = pd.read_csv(input_mapping_file)
    elif country in cfg.GROUP2_COUNTRIES:
        input_mapping_file = cfg.MAPPING_EXP_TAB.format('group2')
        input_mapping = pd.read_csv(input_mapping_file)
    else:
        input_mapping_file = cfg.MAPPING_EXP_TAB.format('other')
        input_mapping = pd.read_csv(input_mapping_file)

    alias_files = [cfg.ALIAS_FILE_INPUT, cfg.ALIAS_FILE_EXP, cfg.ALIAS_FILE_PB, cfg.ALIAS_FILE_AFG,
                   cfg.ALIAS_FILE_IBCM, cfg.ALIAS_FILE_MKTS]
    source_files = [input_source, exp_source, pb_source, afg_source, ibcm_source, mkts_source]
    for alias_file, source_file in zip(alias_files, source_files):
        alias = pd.read_csv(alias_file)
        alias.dropna(how='all', axis=0, inplace=True)
        alias = alias.fillna('')
        if 'start row/col' not in list(alias):
            alias['start row/col'] = ''

        alias_suffix = alias_file[20:-4]
        for idx, row in alias.iterrows():
            if row['Alias'][0] == '#' or row['Alias'] == '':
                continue

            # Check if the alias row is valid
            if row['Alias'][0] in 'rc':
                row['offset'] = 0 if not row['offset'] else row['offset']
                check_alias_row(row, idx, alias_file, source_file)

            if row['Alias'][0] == 'r':
                keyword_row = get_row_index(source_file, row['Keyword'], get_row_index(
                    source_file, row['start row/col']))
                if keyword_row is not None and keyword_row != 'ar':
                    globals()[row['Alias'] + alias_suffix] = keyword_row + int(row['offset'])
            elif row['Alias'][0] == 'c':
                keyword_col = get_col_index(source_file, row['Keyword'], get_col_index(
                    source_file, row['start row/col']))
                if keyword_col is not None and keyword_col != 'ac':
                    globals()[row['Alias'] + alias_suffix] = keyword_col + int(row['offset'])
            elif row['Alias'][0] == 's':
                eval_statement = apply_statement(row['statement'])
                try:
                    globals()[row['Alias']] = eval(str(eval_statement[0][0]))
                except:  # pylint: disable=bare-except
                    logger.error(cfg.MAPPING_ERROR_MESSAGE.format(
                        row['statement'], (idx + 2), alias_file))
                    exit(-1)

    # Process through each mapping and populate values
    for index, row in tqdm(input_mapping.iterrows(), total=input_mapping.shape[0]):
        if row['row_id'][0] == '#':
            continue

        # Check that the aliases are valid
        try:
            row_index = eval(append_suffix(row['row_id'], suffix))
        except NameError:
            logger.error(cfg.INVALID_ALIAS_MESSAGE.format(
                row['row_id'], index + 2, input_mapping_file))
            exit(-1)
        try:
            col_index = eval(append_suffix(row['col_id'], suffix))
        except NameError:
            logger.error(cfg.INVALID_ALIAS_MESSAGE.format(
                row['col_id'], index + 2, input_mapping_file))
            exit(-1)

        # Check that the inputs are valid
        if row_index is None or col_index is None:
            logger.error(cfg.INVALID_MAPPING_MESSAGE.format((index + 2)))
            exit(-1)

        try:
            eval_statement = apply_statement(
                row['statement'], int(row['affected_rows']), int(row['affected_cols']))
        except IndexError:
            logger.error(cfg.MAPPING_ERROR_MESSAGE.format(
                row['statement'], index + 2, input_mapping_file))
            exit(-1)

        for row_num in range(eval_statement.shape[0]):
            for col_num in range(eval_statement.shape[1]):
                try:
                    evaluated_value = eval(str(eval_statement[row_num][col_num]))
                except MissingValueError:
                    logger.warning(cfg.MISSING_VALUE_ERROR.format(
                        str(eval_statement[row_num][col_num]), index + 2, input_mapping_file))
                    continue
                except ValueError:
                    logger.error(cfg.INCORRECT_VALUE_ERROR.format(
                        str(eval_statement[row_num][col_num]), index + 2, input_mapping_file))
                    exit(-1)
                except Exception as err_message:  # pylint: disable=broad-except
                    exp_source.at[row_index + row_num, col_index + col_num] = "#VALUE!"
                    logger.error(cfg.MAPPING_ERROR_MESSAGE.format(
                        row['statement'], (index + 2), input_mapping_file))
                    logger.error("Error details: {}".format(err_message))
                    continue

                if evaluated_value == '':
                    logger.warning(cfg.MISSING_VALUE_ERROR.format(
                        str(eval_statement[row_num][col_num]), index + 2, input_mapping_file))
                    continue

                if isinstance(evaluated_value, np.ndarray):
                    for i in range(0, len(evaluated_value)):
                        for j in range(0, len(evaluated_value[0])):
                            exp_source.at[row_index + row_num + i, col_index + col_num + j] =\
                                evaluated_value.item((i, j))
                else:
                    exp_source.at[row_index + row_num, col_index + col_num] = evaluated_value

    country_report_data['Exp'] = strip_metadata(exp_source)
    # Data frame need to be reshaped before writing to sheet
    rows = dataframe_to_rows(exp_source, index=False, header=False)

    # Write the information back to sheet
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            cell = exp_sheet.cell(row=r_idx, column=c_idx)
            if type(cell).__name__ != 'MergedCell':
                exp_sheet.cell(row=r_idx, column=c_idx, value=value)